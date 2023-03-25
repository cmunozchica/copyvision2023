from django.shortcuts import render, redirect
from django.template import RequestContext
from .models import Trabajo, Producto
from .forms import TrabajoForm, ProductoForm, ProductoFormUpdate, CustomCreationForm
from django.db.models import Q
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.forms import AuthenticationForm 
from django.contrib.auth import login, authenticate, logout, login as auth_login
from django.contrib.auth.decorators import login_required, permission_required
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.template.loader import render_to_string
from weasyprint import HTML
#from weasyprint.fonts import FontConfiguration
from django_weasyprint import WeasyTemplateResponseMixin
from django_weasyprint.views import WeasyTemplateResponse
from django_weasyprint.utils import django_url_fetcher
from .utils import render_to_pdf
from django.views.generic import View
from tablib import Dataset
from .resource import TrabajoResource
from openpyxl import Workbook
from django.http.response import HttpResponse
from django.views.generic import TemplateView
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from django.http import HttpResponseRedirect, HttpResponse, FileResponse
from django.contrib import messages
from django.http import Http404


@permission_required('modulo.add_user')
def UserCreation(request):
    data = {
        'form': CustomCreationForm()
    }
    if request.method == "POST":
        form = CustomCreationForm(data=request.POST)
        if form.is_valid():           
            form.save()
            user = authenticate(username=form.cleaned_data["username"], password=form.cleaned_data["password1"])
            login(request, user)
            messages.success(request, "Usuario registrado correctamente")         
            return redirect('home')
        data["form"] = form

    return render(request, 'registration/UserCreation.html', data )




# Create your views here.
def home(resquest):
    return render(resquest, 'home.html')

def index(resquest):
    return render(resquest, 'index.html')

@permission_required('modulo.view_trabajos')
def trabajoList(request):
    trabajo_queryset = Trabajo.objects.all()
    context = {'trabajo_context': trabajo_queryset,
               'count': trabajo_queryset.count(), }
    return render(request, "trabajo-list.html", context)

@permission_required('modulo.view_trabajos')
def trabajoListGranformato(request):
    trabajo_queryset = Trabajo.objects.filter(tipoTrabajo='GRAN-FORMATO')
    context = {'trabajo_context': trabajo_queryset,
               'count': trabajo_queryset.count(), }
    return render(request, "trabajo-list-granformato.html", context)

    return render(request, "trabajo-list.html", {'trabajo': trabajo})

@permission_required('modulo.view_trabajos')
def trabajoListCorte(request):
    trabajo_queryset = Trabajo.objects.filter(
        tipoTrabajo='CORTE-LASER Y GRABADO')
    context = {'trabajo_context': trabajo_queryset,
               'count': trabajo_queryset.count(), }
    return render(request, "trabajo-list-corte.html", context)

    return render(request, "trabajo-list-corte.html", {'trabajo': trabajo})

@permission_required('modulo.view_trabajos')
def trabajoListImpresiones(request):
    trabajo_queryset = Trabajo.objects.filter(tipoTrabajo='IMPRESIONES')
    context = {'trabajo_context': trabajo_queryset,
               'count': trabajo_queryset.count(), }
    return render(request, "trabajo-list-impresiones.html", context)

    return render(request, "trabajo-list-impresiones.html", {'trabajo': trabajo})

@permission_required('modulo.view_trabajos')
def trabajoListPublicidad(request):
    trabajo_queryset = Trabajo.objects.filter(tipoTrabajo='PUBLICIDAD')
    context = {'trabajo_context': trabajo_queryset,
               'count': trabajo_queryset.count(), }
    return render(request, "trabajo-list-publicidad.html", context)

    return render(request, "trabajo-list-publicidad.html", {'trabajo': trabajo})


@permission_required('modulo.add_trabajo')
def trabajoCreate(request):
    if request.method == "POST":
        form = TrabajoForm(request.POST)
        if form.is_valid():
            try:
                form.save()
                model = form.instance
                return redirect('trabajo-list')
            except:
                pass
    else:
        form = TrabajoForm()
    return render(request, 'trabajo-create.html', {'form': form})

@permission_required('modulo.change_trabajo')
def trabajoUpdate(request, id):
    trabajo = Trabajo.objects.get(id=id)
    form = TrabajoForm(initial={'nit': trabajo.nit, 'cliente': trabajo.cliente, 'direccion': trabajo.direccion, 'telefono': trabajo.telefono, 'fecha': trabajo.fecha,
                       'fechaEntrega': trabajo.fechaEntrega, 'descripcion': trabajo.descripcion, 'tipo': trabajo.tipo, 'tipoTrabajo': trabajo.tipoTrabajo, 'estado': trabajo.estado})
    if request.method == "POST":
        form = TrabajoForm(request.POST, instance=trabajo)
        if form.is_valid():
            try:
                form.save()
                model = form.instance
                return redirect('/trabajo-list')
            except Exception as e:
                pass
    return render(request, 'trabajo-update.html', {'form': form})


def trabajo_detail(request, id):

    trabajo = Trabajo.objects.get(id=id)
    form = TrabajoForm(initial={'nit': trabajo.nit, 'cliente': trabajo.cliente, 'direccion': trabajo.direccion, 'telefono': trabajo.telefono,
                       'fechaEntrega': trabajo.fechaEntrega, 'descripcion': trabajo.descripcion, 'tipo': trabajo.tipo, 'tipoTrabajo': trabajo.tipoTrabajo, 'estado': trabajo.estado})
    return render(request, 'trabajo-detail.html', {'form': form})


class ListTrabajosPdf(View):

    def get(self, request, *args, **kwargs):
        trabajo_queryset = Trabajo.objects.all()
        context = {
            'trabajo_context': trabajo_queryset, 'count': trabajo_queryset.count(),
        }
        pdf = render_to_pdf('report-pdf.html', context)
        return HttpResponse(pdf, content_type='application/pdf')


class ListTrabajosGFPdf(View):

    def get(self, request, *args, **kwargs):
        trabajo_queryset = Trabajo.objects.filter(tipoTrabajo='GRAN-FORMATO')
        context = {
            'trabajo_context': trabajo_queryset, 'count': trabajo_queryset.count(),
        }
        pdf = render_to_pdf('report-pdf.html', context)
        return HttpResponse(pdf, content_type='application/pdf')


class ListTrabajosCortePdf(View):

    def get(self, request, *args, **kwargs):
        trabajo_queryset = Trabajo.objects.filter(
            tipoTrabajo='CORTE-LASER Y GRABADO')
        context = {
            'trabajo_context': trabajo_queryset, 'count': trabajo_queryset.count(),
        }
        pdf = render_to_pdf('report-pdf.html', context)
        return HttpResponse(pdf, content_type='application/pdf')


class ListTrabajosImpPdf(View):

    def get(self, request, *args, **kwargs):
        trabajo_queryset = Trabajo.objects.filter(tipoTrabajo='IMPRESIONES')
        context = {
            'trabajo_context': trabajo_queryset, 'count': trabajo_queryset.count(),
        }
        pdf = render_to_pdf('report-pdf.html', context)
        return HttpResponse(pdf, content_type='application/pdf')


class ListTrabajosPPdf(View):

    def get(self, request, *args, **kwargs):
        trabajo_queryset = Trabajo.objects.filter(tipoTrabajo='PUBLICIDAD')
        context = {
            'trabajo_context': trabajo_queryset, 'count': trabajo_queryset.count(),
        }
        pdf = render_to_pdf('report-pdf.html', context)
        return HttpResponse(pdf, content_type='application/pdf')

@permission_required('modulo.add_producto')
def productoCreate(request):
    if request.method == "POST":
        form = ProductoForm(request.POST)
        if form.is_valid():
            try:
                form.save()
                model = form.instance
                return redirect('producto-list')
            except:
                pass
    else:
        form = ProductoForm()
    return render(request, 'producto-create.html', {'form': form})

@permission_required('modulo.view-producto')
def ProductoList(request):
    busqueda = request.GET.get("buscar")
    producto_queryset = Producto.objects.all()
    page = request.GET.get('page', 1)

    if busqueda:
        producto_queryset = Producto.objects.filter(
            Q(id__icontains=busqueda) |
            Q(nombre__icontains=busqueda) |
            Q(categorias__icontains=busqueda)
        ).distinct()

    try:        
        paginator = Paginator(producto_queryset, 2)    
        producto_queryset = paginator.page(page)
    except:
        raise Http404


    context = {'entity': producto_queryset, 'paginator': paginator}
    return render(request, "producto-list.html", context)


def productoUpdate(request, id):
    producto = Producto.objects.get(id=id)
    form = ProductoFormUpdate(initial={'nombre': producto.nombre, 'precio': producto.precio,
                              'cantidad': producto.cantidad, 'disponibilidad': producto.disponibilidad, 'update': producto.update})
    if request.method == "POST":
        form = ProductoFormUpdate(request.POST, instance=producto)
        if form.is_valid():
            try:
                form.save()
                model = form.instance
                return redirect('/producto-list')
            except Exception as e:
                pass
    return render(request, 'producto-update.html', {'form': form})


def producto_detail(request, id):

    producto = Producto.objects.get(id=id)
    form = ProductoFormUpdate(initial={'nombre': producto.nombre, 'precio': producto.precio,
                              'cantidad': producto.cantidad, 'disponibilidad': producto.disponibilidad, 'update': producto.update})
    return render(request, 'producto-detail.html', {'form': form})


class ListProductosPdf(View):

    def get(self, request, *args, **kwargs):
        producto_queryset = Producto.objects.all()
        context = {
            'producto_context': producto_queryset, 'count': producto_queryset.count(),
        }
        pdf = render_to_pdf('reporte-producto-pdf.html', context)
        return HttpResponse(pdf, content_type='application/pdf')


def importar(request):
    #template = loader.get_template('export/importar.html')
    if request.method == 'POST':
        trabajo_resource = TrabajoResource()
        dataset = Dataset()
        # print(dataset)
        nuevas_trabajo = request.FILES['xlsfile']
        # print(nuevas_personas)
        imported_data = dataset.load(nuevas_trabajo.read())
        # print(dataset)
        result = trabajo_resource.import_data(
            dataset, dry_run=True)  # Test the data import
        # print(result.has_errors())
        if not result.has_errors():
            trabajo_resource.import_data(
                dataset, dry_run=False)  # Actually import now
    return render(request, 'importar.html')


class reporteTrabajoExcel(TemplateView):
    def get(self, request, *args, **kwargs):
        trabajos = Trabajo.objects.all()
        wb = Workbook()
        ws = wb.active
        ws['B1'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B1'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['B1'].fill = PatternFill(
            start_color='66FFCC', end_color='66FFCC', fill_type="solid")
        ws['B1'].font = Font(name='Calibri', size=12, bold=True)
        ws['B1'] = 'REPORTE DE TRABAJOS'
        ws.merge_cells('B1:H1')
        ws['B3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['B3'].fill = PatternFill(
            start_color='66CFCC', end_color='66CFCC', fill_type="solid")
        ws['B3'].font = Font(name='Calibro', size=10, bold=True)
        ws['B3'] = 'CLIENTE'
        ws['C3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['C3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['C3'].fill = PatternFill(
            start_color='66CFCC', end_color='66CFCC', fill_type="solid")
        ws['C3'].font = Font(name='Calibro', size=10, bold=True)
        ws['C3'] = 'FECHA'
        ws['D3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['D3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['D3'].fill = PatternFill(
            start_color='66CFCC', end_color='66CFCC', fill_type="solid")
        ws['D3'].font = Font(name='Calibro', size=10, bold=True)
        ws['D3'] = 'FECHA DE ENTREGA'
        ws['E3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['E3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['E3'].fill = PatternFill(
            start_color='66CFCC', end_color='66CFCC', fill_type="solid")
        ws['E3'].font = Font(name='Calibro', size=10, bold=True)
        ws['E3'] = 'DESCRIPCION'
        ws['F3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['F3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['F3'].fill = PatternFill(
            start_color='66CFCC', end_color='66CFCC', fill_type="solid")
        ws['F3'].font = Font(name='Calibro', size=10, bold=True)
        ws['F3'] = 'TIPO'
        ws['G3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['G3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['G3'].fill = PatternFill(
            start_color='66CFCC', end_color='66CFCC', fill_type="solid")
        ws['G3'].font = Font(name='Calibro', size=10, bold=True)
        ws['G3'] = 'TRABAJO'
        ws['H3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['H3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['H3'].fill = PatternFill(
            start_color='66CFCC', end_color='66CFCC', fill_type="solid")
        ws['H3'].font = Font(name='Calibro', size=10, bold=True)
        ws['H3'] = 'ESTADO'

        cont = 7

        for trabajo in trabajos:
            ws.cell(row=cont, column=2).alignment = Alignment(
                horizontal="center")
            ws.cell(row=cont, column=2).border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                                        top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws.cell(row=cont, column=2).font = Font(name='Calibri', size=8)
            ws.cell(row=cont, column=2).value = trabajo.cliente
            ws.cell(row=cont, column=5).alignment = Alignment(
                horizontal="left")
            ws.cell(row=cont, column=5).border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                                        top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws.cell(row=cont, column=5).font = Font(name='Calibri', size=8)
            ws.cell(row=cont, column=5).value = trabajo.descripcion
            ws.cell(row=cont, column=6).alignment = Alignment(
                horizontal="center")
            ws.cell(row=cont, column=6).border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                                        top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws.cell(row=cont, column=6).font = Font(name='Calibri', size=8)
            ws.cell(row=cont, column=6).value = trabajo.tipo
            ws.cell(row=cont, column=7).alignment = Alignment(
                horizontal="center")
            ws.cell(row=cont, column=7).border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                                        top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws.cell(row=cont, column=7).font = Font(name='Calibri', size=8)
            ws.cell(row=cont, column=7).value = trabajo.tipoTrabajo
            ws.cell(row=cont, column=8).alignment = Alignment(
                horizontal="center")
            ws.cell(row=cont, column=8).border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                                        top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws.cell(row=cont, column=8).font = Font(name='Calibri', size=8)
            ws.cell(row=cont, column=8).value = trabajo.estado
            cont += 1

        nombre_archivo = "ReporteTrabajosExcel.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response


class reporteProductoExcel(TemplateView):
    def get(self, request, *args, **kwargs):
        productos = Producto.objects.all()
        wb = Workbook()
